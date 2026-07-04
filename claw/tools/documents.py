"""Document reading tools: Excel, CSV, PDF, Word — over workspace files.

Parsing libraries are optional (dependency group 'documents'); a tool returns a
clear error if its library isn't installed rather than crashing the agent.
"""

import csv
import io
from pathlib import Path
from typing import Any

from claw.tools.base import Tool

_MAX_CHARS = 40_000


class _WorkspaceDocTool(Tool):
    def __init__(self, workspace: Path):
        self.workspace = workspace.resolve()

    def _resolve(self, raw: str) -> Path | None:
        try:
            p = (self.workspace / raw).resolve() if not Path(raw).is_absolute() else Path(raw).resolve()
            p.relative_to(self.workspace)
        except (ValueError, OSError):
            return None
        return p if p.is_file() else None

    @staticmethod
    def _cap(text: str) -> str:
        return text if len(text) <= _MAX_CHARS else text[:_MAX_CHARS] + "\n... (truncated)"


class ReadExcelTool(_WorkspaceDocTool):
    name = "read_excel"
    description = "Read an Excel workbook (.xlsx) from the workspace; returns each sheet as rows."
    parameters = {
        "type": "object",
        "properties": {
            "path": {"type": "string"},
            "sheet": {"type": "string", "description": "Optional sheet name; default all"},
            "max_rows": {"type": "integer", "description": "Rows per sheet (default 200)"},
        },
        "required": ["path"],
    }

    async def execute(self, path: str, sheet: str = "", max_rows: int = 200, **_: Any) -> str:
        target = self._resolve(path)
        if target is None:
            return f"Error: file not found: {path}"
        try:
            from openpyxl import load_workbook
        except ImportError:
            return "Error: Excel support is not installed (pip install openpyxl)."
        wb = load_workbook(target, read_only=True, data_only=True)
        names = [sheet] if sheet else wb.sheetnames
        out: list[str] = []
        for name in names:
            if name not in wb.sheetnames:
                out.append(f"[sheet '{name}' not found]")
                continue
            ws = wb[name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_rows:
                    rows.append("... (more rows)")
                    break
                rows.append(" | ".join("" if c is None else str(c) for c in row))
            out.append(f"## Sheet: {name}\n" + "\n".join(rows))
        wb.close()
        return self._cap("\n\n".join(out))


class ReadCsvTool(_WorkspaceDocTool):
    name = "read_csv"
    description = "Read a CSV/TSV file from the workspace as delimited rows."
    parameters = {
        "type": "object",
        "properties": {"path": {"type": "string"}, "max_rows": {"type": "integer"}},
        "required": ["path"],
    }

    async def execute(self, path: str, max_rows: int = 500, **_: Any) -> str:
        target = self._resolve(path)
        if target is None:
            return f"Error: file not found: {path}"
        text = target.read_text("utf-8", "replace")
        delimiter = "\t" if target.suffix.lower() == ".tsv" else ","
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        lines = []
        for i, row in enumerate(reader):
            if i >= max_rows:
                lines.append("... (more rows)")
                break
            lines.append(" | ".join(row))
        return self._cap("\n".join(lines))


class ReadPdfTool(_WorkspaceDocTool):
    name = "read_pdf"
    description = "Extract text from a PDF file in the workspace."
    parameters = {"type": "object", "properties": {"path": {"type": "string"}}, "required": ["path"]}

    async def execute(self, path: str, **_: Any) -> str:
        target = self._resolve(path)
        if target is None:
            return f"Error: file not found: {path}"
        try:
            from pypdf import PdfReader
        except ImportError:
            return "Error: PDF support is not installed (pip install pypdf)."
        reader = PdfReader(str(target))
        parts = [(page.extract_text() or "") for page in reader.pages]
        return self._cap("\n\n".join(parts).strip() or "(no extractable text)")


class ReadDocxTool(_WorkspaceDocTool):
    name = "read_docx"
    description = "Extract text from a Word (.docx) file in the workspace."
    parameters = {"type": "object", "properties": {"path": {"type": "string"}}, "required": ["path"]}

    async def execute(self, path: str, **_: Any) -> str:
        target = self._resolve(path)
        if target is None:
            return f"Error: file not found: {path}"
        try:
            import docx
        except ImportError:
            return "Error: Word support is not installed (pip install python-docx)."
        document = docx.Document(str(target))
        return self._cap("\n".join(p.text for p in document.paragraphs).strip() or "(empty document)")


def build_document_tools(workspace: Path) -> list[Tool]:
    return [
        ReadExcelTool(workspace),
        ReadCsvTool(workspace),
        ReadPdfTool(workspace),
        ReadDocxTool(workspace),
    ]
